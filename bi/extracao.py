############ Análise de dados educacionais - TCC Fernanda e Luiz ############
# Autores: Fernanda Amaral Melo e Luiz Fernando Araújo 
# Contato: fernanda.amaral.melo@gmail.com e luizfna@gmail.com

# Script usado para extrair os dados dos arquivos .xlsx e .csv salvos na pasta
# "data" (provenientes da plataforma de BI da Universidade de Brasília) para 
# inserção em um banco de dados SQL

import openpyxl
import numpy as np
from openpyxl import Workbook
import psycopg2
import json

with open('credentials.json') as json_file:
    data = json.load(json_file)

conn = psycopg2.connect(
    host=data['host'],
    port=data['port'],
    user=data['user'],
    password=data['password']
)

cur = conn.cursor()
cur.execute("CREATE TABLE IF NOT EXISTS admission_vs_dropout_mode (id serial PRIMARY KEY, student_id TEXT, dropout_year INTEGER, dropout_semester INTEGER, admission_year INTEGER, admission_semester INTEGER, dropout_mode VARCHAR);")

book = openpyxl.load_workbook('data/ingresso_saida_anonimo.xlsx', data_only=True)
sheet = book.active
rows = sheet.rows

values = []
last_period = ""

for table_row in range(4, len(sheet['B']) + 1):
    student_id = sheet.cell(row=table_row, column=2).value
    if(sheet.cell(row=table_row, column=1).value != None):
        last_period = sheet.cell(row=table_row, column=1).value
    admission_period = last_period

    dropout_mode = ""
    for table_column in range(3, len(sheet[table_row]) + 1):
        if(sheet.cell(row=1, column=table_column).value):
            dropout_mode = sheet.cell(row=1, column=table_column).value
        if(sheet.cell(row=table_row, column=table_column).value):
            dropout_period = sheet.cell(row=2, column=table_column).value
            break
        else:
            dropout_period="0/0"

    if(dropout_period == "Ativo"):  
        cur.execute("INSERT INTO admission_vs_dropout_mode (student_id, dropout_year, dropout_semester, admission_year, admission_semester, dropout_mode) VALUES (%s, NULL, NULL, %s, %s, %s)", (student_id, int(admission_period.split('/')[0]), int(admission_period.split('/')[1]), dropout_mode))
    else:    
        cur.execute("INSERT INTO admission_vs_dropout_mode (student_id, dropout_year, dropout_semester, admission_year, admission_semester, dropout_mode) VALUES (%s, %s, %s, %s, %s, %s)", (student_id, int(dropout_period.split('/')[0]), int(dropout_period.split('/')[1]), int(admission_period.split('/')[0]), int(admission_period.split('/')[1]), dropout_mode))

cur.execute("SELECT * FROM admission_vs_dropout_mode;")      
result = cur.fetchall()
print("Done")
conn.commit()
cur.close()
conn.close()