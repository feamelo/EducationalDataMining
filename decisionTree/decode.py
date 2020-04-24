############ Análise de dados educacionais - TCC Fernanda e Luiz ############
# Autores: Fernanda Amaral Melo e Luiz Fernando Araújo 
# Contato: fernanda.amaral.melo@gmail.com e luizfna@gmail.com

# Script usado para extrair os dados dos arquivos .xlsx e .csv salvos na pasta
# "data" (provenientes da plataforma de BI da Universidade de Brasília) para 
# inserção em um banco de dados SQL

import openpyxl
import numpy as np
from openpyxl import Workbook
import psycopg2
import subprocess
import sys

with open('credentials.json') as json_file:
    data = json.load(json_file)

conn = psycopg2.connect(
    host=data['host'],
    port=data['port'],
    dbname=data['dbname'],
    user=data['user'],
    password=data['password']
)

cur = conn.cursor()
cur.execute("CREATE TABLE IF NOT EXISTS codes (id serial PRIMARY KEY, code INTEGER, name TEXT);")

book = openpyxl.load_workbook('data/courses_codes.xlsx', data_only=True)
sheet = book.active
rows = sheet.rows

values = []
last_period = ""

for table_row in range(1, len(sheet['A']) + 1):
    code = sheet.cell(row=table_row, column=1).value
    name = sheet.cell(row=table_row, column=2).value
    if(type(code) is int):
        cur.execute("INSERT INTO codes (code, name) VALUES (%s, %s)", (code, name))

print("Done")
conn.commit()
cur.close()
conn.close()